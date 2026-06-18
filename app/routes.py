import csv
import io
import json
import os
import platform
import subprocess
import shutil
import tempfile
from calendar import Calendar, month_name
from datetime import date, datetime, timedelta
from zipfile import ZipFile, ZIP_DEFLATED

from flask import Blueprint, render_template, redirect, url_for, flash, request, Response, send_file, current_app, abort
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash

from .models import db, User, Settings, Category, RecurringBill, BillOccurrence, PlannedPurchase, AccountBalanceSnapshot, IncomeSource, SharedIncomeAllocation, Bucket, DashboardWidget, PaydayChecklistItem, PaydayChecklistPreference, AuditLog, NotificationSetting, CycleCloseout
from .forms import LoginForm, UserForm, UserProfileForm, SettingsForm, CategoryForm, RecurringBillForm, PlannedPurchaseForm, AccountBalanceForm, IncomeSourceForm, SharedIncomeAllocationForm, BucketForm, NotificationSettingsForm, CycleCloseoutForm
from .version import APP_VERSION, APP_RELEASE_NAME
from .budget_engine import (
    annual_cost, fortnightly_bill_amount, generate_bill_dates,
    planned_purchase_fortnightly_amount, current_pay_cycle, household_pay_cycle, money, parse_date, income_for_cycle,
    generate_income_dates, next_income_pay_date, calculate_bucket_allocations, calculate_person_bucket_allocations,
    is_shared_purchase, planned_purchase_scope_label, calculate_shared_income_bucket_additions,
    apply_shared_income_allocations,
)

main = Blueprint("main", __name__)

MAX_IMPORT_PREVIEW_ROWS = 250
MAX_IMPORT_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_RESTORE_UPLOAD_BYTES = 50 * 1024 * 1024


def check_upload_size(file_storage, max_bytes):
    """Return False when an uploaded file exceeds a configured byte limit."""
    if not file_storage:
        return False
    stream = file_storage.stream
    current = stream.tell()
    stream.seek(0, os.SEEK_END)
    size = stream.tell()
    stream.seek(current)
    return size <= max_bytes


def bill_import_preview_path():
    return os.path.join(current_app.instance_path, "bill_import_preview.json")


def save_bill_import_preview(rows):
    os.makedirs(current_app.instance_path, exist_ok=True)
    with open(bill_import_preview_path(), "w", encoding="utf-8") as preview_file:
        json.dump(rows, preview_file)


def load_bill_import_preview():
    path = bill_import_preview_path()
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as preview_file:
            return json.load(preview_file)
    except (OSError, json.JSONDecodeError):
        return []


def clear_bill_import_preview():
    try:
        os.unlink(bill_import_preview_path())
    except FileNotFoundError:
        pass


def send_temp_file(path, *, download_name, mimetype=None):
    """Send a generated temporary file and remove it when the response closes."""
    response = send_file(path, as_attachment=True, download_name=download_name, mimetype=mimetype)

    def cleanup():
        try:
            os.unlink(path)
        except OSError:
            pass

    response.call_on_close(cleanup)
    return response


def get_settings():
    return Settings.query.first()


def category_choices(kind):
    """Return category choices for bill or purchase forms."""
    rows = Category.query.filter(Category.active.is_(True)).order_by(Category.name).all()
    choices = [(0, "Uncategorised")]
    for category in rows:
        if category.category_type in [kind, "Both"]:
            choices.append((category.id, category.name))
    return choices


def regenerate_bill_occurrences(bill, scope="future_unpaid"):
    """Regenerate unpaid occurrences for a recurring bill.

    Paid occurrences are kept so payment history is not destroyed. By default,
    only future unpaid occurrences are replaced. This makes editing safer after
    the app has been used for a while. The old behaviour can still be requested
    by using scope="all_unpaid".
    """
    settings = get_settings()
    budget_year = settings.budget_year
    today_iso = date.today().isoformat()

    query = BillOccurrence.query.filter(
        BillOccurrence.recurring_bill_id == bill.id,
        BillOccurrence.status != "Paid",
    )
    if scope == "future_unpaid":
        query = query.filter(BillOccurrence.due_date >= today_iso)
    query.delete(synchronize_session=False)

    for due_date in generate_bill_dates(bill, budget_year):
        due_iso = due_date.isoformat()
        if scope == "future_unpaid" and due_iso < today_iso:
            continue
        existing = BillOccurrence.query.filter_by(
            recurring_bill_id=bill.id,
            due_date=due_iso,
        ).first()
        if not existing:
            db.session.add(BillOccurrence(
                recurring_bill_id=bill.id,
                due_date=due_iso,
                amount=bill.amount,
                status="Upcoming",
            ))



def enforce_single_remainder_bucket(preferred_bucket=None):
    """Ensure only one bucket has the remainder-cap option enabled.

    The pay split calculation is ordered, so allowing multiple remainder buckets
    would make the later ones confusing. When a preferred bucket is provided, it
    becomes the only capped bucket; otherwise the first sorted capped bucket is kept.
    """
    capped = Bucket.query.filter(Bucket.cap_to_remaining.is_(True)).order_by(Bucket.sort_order, Bucket.name).all()
    if not capped:
        return
    keep = preferred_bucket if preferred_bucket and preferred_bucket.cap_to_remaining else capped[0]
    for bucket in capped:
        if bucket.id != keep.id:
            bucket.cap_to_remaining = False




def prepare_bucket_form_for_display(form, bucket=None):
    """Set the bucket form's amount type based on saved values.

    Buckets can use either a percentage or a fixed household amount. The form
    keeps this explicit so users do not accidentally enter both.
    """
    if bucket and bucket.fixed_amount is not None:
        form.allocation_method.data = "fixed"
    elif not form.allocation_method.data:
        form.allocation_method.data = "percentage"


def apply_bucket_form(bucket, form):
    """Apply bucket form data while enforcing one amount type only."""
    bucket.name = form.name.data.strip()
    bucket.allocation_method = getattr(form, "allocation_method", None) and form.allocation_method.data

    if form.allocation_method.data == "fixed":
        fixed = form.fixed_amount.data
        if fixed is None or fixed <= 0:
            form.fixed_amount.errors.append("Enter a fixed amount greater than $0.")
            return False
        bucket.fixed_amount = money(fixed)
        bucket.percentage = 0
    else:
        percentage = form.percentage.data
        if percentage is None or percentage < 0:
            form.percentage.errors.append("Enter a percentage of 0 or higher.")
            return False
        bucket.percentage = money(percentage)
        bucket.fixed_amount = None

    bucket.rounding_increment = form.rounding_increment.data
    bucket.cap_to_remaining = bool(form.cap_to_remaining.data)
    bucket.bucket_type = form.bucket_type.data
    bucket.active = bool(form.active.data)
    bucket.sort_order = form.sort_order.data or 0
    bucket.notes = form.notes.data
    return True

def normalise_date_string(value):
    parsed = parse_date(value)
    return parsed.isoformat() if parsed else None


def apply_bill_form(form, bill):
    """Apply the recurring bill form to the model using one clear first due date.

    Internally Solace still stores due_day, due_month, and start_date because the
    recurrence engine uses those fields. The form presents a simpler model: the
    user enters the first date the bill comes out, and Solace derives the rest.
    """
    first_due = parse_date(form.first_due_date.data)
    if not first_due:
        raise ValueError("A first due date is required.")

    bill.name = form.name.data
    bill.amount = form.amount.data
    bill.frequency = form.frequency.data
    bill.due_day = first_due.day
    bill.start_date = first_due.isoformat()

    # Month-based recurring items need an anchor month. Monthly bills do not,
    # but quarterly/six-monthly/yearly bills repeat from the first due month.
    if bill.frequency in ["Quarterly", "Six-monthly", "Yearly"]:
        bill.due_month = first_due.month
    else:
        bill.due_month = None

    bill.end_date = normalise_date_string(form.end_date.data) if form.end_date.data else None
    bill.active = bool(form.active.data)
    bill.autopay = bool(form.autopay.data)
    bill.account_name = form.account_name.data
    bill.include_in_set_aside = bool(form.include_in_set_aside.data)
    bill.notes = form.notes.data


def get_or_create_category(name, category_type="Bill"):
    if not name:
        return None
    name = str(name).strip()
    if not name:
        return None
    category = Category.query.filter(db.func.lower(Category.name) == name.lower()).first()
    if category:
        return category
    category = Category(name=name, category_type=category_type, active=True)
    db.session.add(category)
    db.session.flush()
    return category



def audit(action, entity_type=None, entity_name=None, details=None):
    """Record a lightweight audit event for important user actions."""
    try:
        db.session.add(AuditLog(
            created_at=datetime.now().isoformat(timespec="seconds"),
            action=action,
            entity_type=entity_type,
            entity_name=entity_name,
            details=details,
        ))
    except Exception:
        # Audit logging should never block the main workflow.
        pass


def get_database_path():
    """Return the SQLite database path from the configured SQLAlchemy URI."""
    uri = current_app.config["SQLALCHEMY_DATABASE_URI"]
    return uri.replace("sqlite:///", "") if uri.startswith("sqlite:///") else uri


def get_git_commit():
    """Return the current Git commit when the source checkout is available."""
    env_commit = os.environ.get("SOLACE_GIT_COMMIT")
    if env_commit:
        return env_commit[:12]
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=current_app.root_path + "/..",
            check=True,
            capture_output=True,
            text=True,
            timeout=2,
        )
        return result.stdout.strip()
    except Exception:
        return "not available in this build"


def file_size_label(path):
    """Return a compact file-size label for diagnostics."""
    try:
        size = os.path.getsize(path)
    except OSError:
        return "not found"
    units = ["B", "KB", "MB", "GB"]
    value = float(size)
    for unit in units:
        if value < 1024 or unit == units[-1]:
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024


def parse_bill_import_rows(rows):
    """Convert uploaded rows into a preview with validation before committing."""
    parsed_rows = []
    errors = 0
    allowed_frequencies = {"Weekly", "Fortnightly", "Monthly", "Quarterly", "Six-monthly", "Yearly"}
    for index, row in enumerate(rows, start=2):
        preview = {"source_row": index, "errors": []}
        try:
            name = str(pick(row, "name", "bill", "bill_name")).strip()
            if not name:
                preview["errors"].append("Missing name")
            amount = float(pick(row, "amount", "cost", default=0))
            if amount <= 0:
                preview["errors"].append("Amount must be greater than 0")
            frequency = str(pick(row, "frequency", default="Monthly")).strip().title()
            if frequency == "Six Monthly":
                frequency = "Six-monthly"
            if frequency not in allowed_frequencies:
                preview["errors"].append(f"Unsupported frequency: {frequency}")
            due_day = int(float(pick(row, "due_day", "day", "due", default=1)))
            if due_day < 1 or due_day > 31:
                preview["errors"].append("Due day must be 1-31")
            due_month_raw = pick(row, "due_month", "month", default="")
            due_month = int(float(due_month_raw)) if due_month_raw not in [None, ""] else None
            if due_month is not None and (due_month < 1 or due_month > 12):
                preview["errors"].append("Due month must be 1-12")
            start_date = normalise_date_string(pick(row, "start_date", "start", default=f"{get_settings().budget_year}-01-01"))
            end_value = pick(row, "end_date", "end", default="")
            end_date = normalise_date_string(end_value) if end_value else None
            category_name = str(pick(row, "category", default="")).strip()
            active = str(pick(row, "active", default="yes")).strip().lower() not in ["no", "false", "0"]
            autopay = str(pick(row, "autopay", "auto_pay", default="no")).strip().lower() in ["yes", "true", "1"]
            include = str(pick(row, "include_in_set_aside", "include", default="yes")).strip().lower() not in ["no", "false", "0"]
            preview.update({
                "name": name,
                "amount": amount,
                "frequency": frequency,
                "due_day": due_day,
                "due_month": due_month,
                "start_date": start_date,
                "end_date": end_date,
                "category": category_name,
                "active": active,
                "autopay": autopay,
                "account_name": str(pick(row, "account_name", "account", default="")).strip(),
                "include_in_set_aside": include,
                "notes": str(pick(row, "notes", default="")).strip(),
            })
        except Exception as exc:
            preview["errors"].append(str(exc))
        if preview["errors"]:
            errors += 1
        parsed_rows.append(preview)
    return parsed_rows, errors


def ensure_payday_checklist_items(settings, cycle_start, income_items, person_bucket_allocations, bucket_allocations, include_hidden=False):
    """Create checklist items for the current pay cycle if they do not already exist.

    Transfer items can be hidden using PaydayChecklistPreference. This is for
    automatic transfers that do not need to appear in the active payday list.
    """
    existing = {item.item_key: item for item in PaydayChecklistItem.query.filter_by(cycle_start=cycle_start.isoformat()).all()}
    hidden_preferences = {pref.item_key for pref in PaydayChecklistPreference.query.filter_by(hidden=True).all()}
    required = []
    order = 10
    required.append(("confirm_income", "Confirm all expected income has arrived", None, order)); order += 10
    for person in person_bucket_allocations:
        for row in person.get("bucket_allocations", []):
            amount = money(row["rounded_amount"])
            if amount <= 0:
                continue
            key = f"transfer_{person['person']}_{row['bucket'].id}".lower().replace(" ", "_")
            label = f"{person['person']}: transfer {settings.currency_symbol}{amount:.2f} to {row['bucket'].name}"
            required.append((key, label, amount, order)); order += 10
    required.append(("review_due_bills", "Review bills due before next payday", None, order)); order += 10
    required.append(("record_balance", "Optional: record the bills/set-aside account balance", None, order))
    for key, label, amount, sort_order in required:
        if key not in existing:
            db.session.add(PaydayChecklistItem(
                cycle_start=cycle_start.isoformat(),
                item_key=key,
                label=label,
                amount=amount,
                completed=False,
                sort_order=sort_order,
            ))
    db.session.commit()
    query = PaydayChecklistItem.query.filter_by(cycle_start=cycle_start.isoformat())
    items = query.order_by(PaydayChecklistItem.sort_order).all()
    if include_hidden:
        return items
    return [item for item in items if item.item_key not in hidden_preferences]

def get_dashboard_widgets():
    """Return dashboard widgets in display order and a quick enabled lookup."""
    widgets = DashboardWidget.query.order_by(DashboardWidget.sort_order, DashboardWidget.title).all()
    if not widgets:
        # Existing databases should normally be seeded at startup. This fallback
        # prevents a blank dashboard if the table exists but has no rows.
        defaults = [
            ("quick_links", "Quick links", True, 8, "small"),
            ("set_aside_summary", "Set-aside summary", True, 10, "wide"),
            ("income_summary", "Income summary", True, 20, "medium"),
            ("bucket_summary", "Bucket summary", True, 30, "medium"),
            ("per_person_contributions", "Individual contributions", True, 40, "wide"),
            ("bills_bucket_health", "Bills bucket health", True, 45, "medium"),
            ("payday_checklist", "Payday checklist", True, 48, "medium"),
            ("due_before_next_payday", "Due this cycle", True, 50, "wide"),
        ]
        for key, title, enabled, sort_order, size in defaults:
            db.session.add(DashboardWidget(widget_key=key, title=title, enabled=enabled, sort_order=sort_order, size=size))
        db.session.commit()
        widgets = DashboardWidget.query.order_by(DashboardWidget.sort_order, DashboardWidget.title).all()
    enabled = {widget.widget_key for widget in widgets if widget.enabled}
    return widgets, enabled




def get_cycle_window(settings, income_sources=None, offset=0, today=None):
    """Return the household pay-cycle window, optionally shifted by one cycle.

    Active income sources are treated as the source of truth. Their stored
    dates are known-payday anchors and can be in the past. This function must
    not call itself; it calculates the current cycle directly from the earliest
    active income anchor and then applies the requested offset.

    Pay frequency (weekly or fortnightly) is read from Settings.pay_frequency.
    """
    today = today or date.today()
    frequency = getattr(settings, "pay_frequency", "fortnightly") or "fortnightly"
    interval = 7 if frequency == "weekly" else 14

    active_sources = [
        source for source in (income_sources or [])
        if getattr(source, "active", False) and getattr(source, "next_pay_date", None)
    ]

    anchors = []
    for source in active_sources:
        anchor = parse_date(getattr(source, "next_pay_date", None))
        if anchor:
            anchors.append(anchor)

    if anchors:
        cycle_start = min(anchors)
    else:
        cycle_start = parse_date(getattr(settings, "first_payday", None)) or today

    while cycle_start + timedelta(days=interval - 1) < today:
        cycle_start += timedelta(days=interval)

    while cycle_start > today:
        cycle_start -= timedelta(days=interval)

    if offset:
        cycle_start += timedelta(days=interval * int(offset))

    cycle_end = cycle_start + timedelta(days=interval - 1)
    next_payday = cycle_end + timedelta(days=1)
    return cycle_start, cycle_end, next_payday
    return cycle_start, cycle_end, next_payday


def cycle_bill_cutoff(settings, cycle_end, next_payday):
    """Return the final bill date included in a pay cycle.

    By default, bills due on payday belong to the new cycle. Some households
    may prefer the previous cycle, so this is configurable in Settings.
    """
    if getattr(settings, "payday_bill_handling", "new_cycle") == "previous_cycle":
        return next_payday
    return cycle_end


def get_cycle_occurrences(settings, cycle_start, cycle_end, next_payday, status=None):
    cutoff = cycle_bill_cutoff(settings, cycle_end, next_payday)
    query = BillOccurrence.query.filter(
        BillOccurrence.due_date >= cycle_start.isoformat(),
        BillOccurrence.due_date <= cutoff.isoformat(),
    )
    if status:
        query = query.filter(BillOccurrence.status == status)
    return query.order_by(BillOccurrence.due_date, BillOccurrence.id).all()


def get_or_create_cycle_closeout(cycle_start, cycle_end):
    closeout = CycleCloseout.query.filter_by(cycle_start=cycle_start.isoformat()).first()
    if not closeout:
        closeout = CycleCloseout(cycle_start=cycle_start.isoformat(), cycle_end=cycle_end.isoformat(), status="Open")
        db.session.add(closeout)
        db.session.flush()
    return closeout


def purchase_person_choices():
    """Return person choices based on configured income owners."""
    names = sorted({
        (source.owner_name or "Household").strip()
        for source in IncomeSource.query.filter_by(active=True).all()
        if (source.owner_name or "").strip()
    }, key=str.lower)
    choices = [("", "Select person")]
    choices.extend((name, name) for name in names)
    return choices


def prepare_purchase_form(form, purchase=None):
    form.category_id.choices = category_choices("Purchase")
    form.owner_name.choices = purchase_person_choices()
    if purchase and request.method == "GET":
        form.category_id.data = purchase.category_id or 0
        form.owner_name.data = purchase.owner_name or ""


def shared_active_purchases():
    return [p for p in PlannedPurchase.query.filter_by(status="Active").all() if is_shared_purchase(p)]


def individual_purchase_contributions(first_payday):
    rows = {}
    for purchase in PlannedPurchase.query.filter_by(status="Active").all():
        if is_shared_purchase(purchase):
            continue
        person = purchase.owner_name or "Unassigned"
        rows.setdefault(person, {"person": person, "purchases": [], "total": 0})
        amount = planned_purchase_fortnightly_amount(purchase, first_payday)
        rows[person]["purchases"].append({"purchase": purchase, "amount": amount})
        rows[person]["total"] = money(rows[person]["total"] + amount)
    return sorted(rows.values(), key=lambda row: row["person"].lower())


def planned_purchase_summary_rows(purchases, first_payday):
    """Build display-ready planned-purchase summary rows."""
    rows = []
    for purchase in purchases:
        target = float(purchase.target_amount or 0)
        saved = float(purchase.amount_saved or 0)
        remaining = max(target - saved, 0)
        progress = min(100, (saved / target * 100) if target else 0)
        rows.append({
            "purchase": purchase,
            "remaining": money(remaining),
            "progress": progress,
            "per_fortnight": money(planned_purchase_fortnightly_amount(purchase, first_payday)),
            "is_shared": is_shared_purchase(purchase),
            "owner": purchase.owner_name or "Unassigned",
        })
    return rows


def planned_purchase_totals(rows):
    """Return totals for a list of planned-purchase summary rows."""
    return {
        "target": money(sum(row["purchase"].target_amount or 0 for row in rows)),
        "saved": money(sum(row["purchase"].amount_saved or 0 for row in rows)),
        "remaining": money(sum(row["remaining"] for row in rows)),
        "per_fortnight": money(sum(row["per_fortnight"] for row in rows if row["purchase"].status == "Active")),
        "active_count": sum(1 for row in rows if row["purchase"].status == "Active"),
    }


@main.context_processor
def inject_globals():
    settings = get_settings()
    theme_mode = settings.theme if settings and settings.theme in ["Light", "Dark", "Auto"] else "Light"
    show_help_tips = True if not settings else bool(getattr(settings, "show_help_tips", True))
    return {
        "settings": settings,
        "money": money,
        "theme_mode": theme_mode,
        "show_help_tips": show_help_tips,
        "planned_purchase_scope_label": planned_purchase_scope_label,
        "app_version": APP_VERSION,
        "app_release_name": APP_RELEASE_NAME,
    }


@main.route("/health")
def health():
    return {"status": "ok", "app": "Project Solace"}


@main.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("main.dashboard"))
    form = LoginForm()
    users = User.query.filter_by(active=True).order_by(User.display_name).all()
    settings = get_settings()
    if form.validate_on_submit():
        try:
            selected_user_id = int(form.selected_user_id.data)
        except (ValueError, TypeError):
            flash("Please select a user first.", "danger")
            return render_template("login.html", form=form, users=users, settings=settings)
        user = db.session.get(User, selected_user_id)
        if not user or not user.active:
            flash("User not found.", "danger")
            return render_template("login.html", form=form, users=users, settings=settings)
        if check_password_hash(user.password_hash, form.password.data):
            login_user(user)
            return redirect(url_for("main.dashboard"))
        flash("Incorrect PIN. Please try again.", "danger")
    return render_template("login.html", form=form, users=users, settings=settings)


@main.route("/users")
@login_required
def manage_users():
    all_users = User.query.order_by(User.display_name).all()
    form = UserForm()
    return render_template("users.html", users=all_users, form=form)


@main.route("/users/add", methods=["POST"])
@login_required
def add_user():
    form = UserForm()
    if form.validate_on_submit():
        if not form.pin.data:
            flash("A PIN is required when creating a new user.", "danger")
            return redirect(url_for("main.manage_users"))
        base_username = form.display_name.data.strip().lower().replace(" ", "_")
        username = base_username
        counter = 1
        while User.query.filter_by(username=username).first():
            username = f"{base_username}_{counter}"
            counter += 1
        user = User(
            username=username,
            display_name=form.display_name.data.strip(),
            avatar_emoji=form.avatar_emoji.data.strip() or "🏠",
            active=form.active.data,
            role="admin",
        )
        user.set_password(form.pin.data)
        db.session.add(user)
        db.session.commit()
        flash(f"User '{user.display_name}' created.", "success")
    return redirect(url_for("main.manage_users"))


@main.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
def edit_user(user_id):
    user = db.session.get(User, user_id)
    if not user:
        abort(404)
    form = UserForm(obj=user)
    if request.method == "GET":
        form.pin.data = ""
    if form.validate_on_submit():
        user.display_name = form.display_name.data.strip()
        user.avatar_emoji = form.avatar_emoji.data.strip() or "🏠"
        user.active = form.active.data
        if form.pin.data:
            user.set_password(form.pin.data)
        db.session.commit()
        flash(f"User '{user.display_name}' updated.", "success")
        return redirect(url_for("main.manage_users"))
    return render_template("user_edit.html", user=user, form=form)


@main.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
def delete_user(user_id):
    user = db.session.get(User, user_id)
    if not user:
        abort(404)
    if User.query.count() <= 1:
        flash("Cannot delete the only user account.", "danger")
        return redirect(url_for("main.manage_users"))
    db.session.delete(user)
    db.session.commit()
    flash("User deleted.", "success")
    return redirect(url_for("main.manage_users"))


@main.route("/account", methods=["GET", "POST"])
@login_required
def account():
    form = UserProfileForm(obj=current_user)
    if request.method == "GET":
        form.current_pin.data = ""
        form.new_pin.data = ""
    if form.validate_on_submit():
        current_user.display_name = form.display_name.data.strip()
        current_user.avatar_emoji = form.avatar_emoji.data.strip() or "🏠"
        if form.new_pin.data:
            if not form.current_pin.data or not check_password_hash(current_user.password_hash, form.current_pin.data):
                flash("Current PIN is incorrect. PIN not changed.", "danger")
                return render_template("account.html", form=form)
            current_user.set_password(form.new_pin.data)
        db.session.commit()
        flash("Account updated.", "success")
        return redirect(url_for("main.account"))
    return render_template("account.html", form=form)


@main.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("main.login"))


@main.route("/")
@login_required
def dashboard():
    settings = get_settings()
    bills = RecurringBill.query.filter_by(active=True).order_by(RecurringBill.name).all()
    purchases = PlannedPurchase.query.filter_by(status="Active").order_by(PlannedPurchase.target_date).all()
    shared_purchases = [p for p in purchases if is_shared_purchase(p)]
    individual_purchase_rows = individual_purchase_contributions(settings.first_payday)

    bill_fortnightly_total = money(sum(fortnightly_bill_amount(b) for b in bills))
    purchase_fortnightly_total = money(sum(planned_purchase_fortnightly_amount(p, settings.first_payday) for p in shared_purchases))
    buffer_amount = money(settings.default_buffer_amount)
    total_set_aside = money(bill_fortnightly_total + purchase_fortnightly_total + buffer_amount)

    today = date.today()
    income_sources = IncomeSource.query.filter_by(active=True).order_by(IncomeSource.next_pay_date, IncomeSource.name).all()
    cycle_start, cycle_end, next_payday = get_cycle_window(settings, income_sources, today=today)
    income_items, income_total = income_for_cycle(income_sources, cycle_start, cycle_end)

    # Separate individual and shared income. Bucket percentage math runs only
    # against individual income so shared income doesn't inflate personal splits.
    individual_income_total = money(sum(
        item["amount"] for item in income_items
        if getattr(item["source"], "income_scope", "Individual") != "Shared"
    ))
    shared_income_total = money(income_total - individual_income_total)

    buckets = Bucket.query.filter_by(active=True).order_by(Bucket.sort_order, Bucket.name).all()

    # Standard bucket math runs against individual income only.
    bucket_allocations = calculate_bucket_allocations(buckets, individual_income_total)
    person_bucket_allocations = calculate_person_bucket_allocations(income_items, buckets, individual_income_total)

    # Apply shared income on top: standard-mode shared income adds to the pool
    # after per-person splits; lump/custom go directly to nominated buckets.
    shared_bucket_additions, shared_standard_pool = calculate_shared_income_bucket_additions(income_items, buckets)
    if shared_standard_pool:
        standard_shared_allocations = calculate_bucket_allocations(buckets, shared_standard_pool)
        for i, row in enumerate(bucket_allocations):
            row["rounded_amount"] = money(row["rounded_amount"] + standard_shared_allocations[i]["rounded_amount"])
            row["raw_amount"] = money(row["raw_amount"] + standard_shared_allocations[i]["raw_amount"])
    for row in bucket_allocations:
        extra = shared_bucket_additions.get(row["bucket"].id, 0)
        if extra:
            row["rounded_amount"] = money(row["rounded_amount"] + extra)
            row["raw_amount"] = money(row["raw_amount"] + extra)

    total_bucket_amount = money(sum(row["rounded_amount"] for row in bucket_allocations))
    remaining_after_buckets = money(income_total - total_bucket_amount)
    bills_bucket_total = money(sum(row["rounded_amount"] for row in bucket_allocations if row["bucket"].bucket_type == "Bills"))
    bills_and_purchases_bucket_total = money(sum(row["rounded_amount"] for row in bucket_allocations if row["bucket"].bucket_type in ["Bills", "Planned purchases"]))
    bills_bucket_delta = money(bills_bucket_total - bill_fortnightly_total)
    bills_bucket_status = "covered" if bills_bucket_delta >= 0 else "short"

    checklist_items = ensure_payday_checklist_items(settings, cycle_start, income_items, person_bucket_allocations, bucket_allocations)
    checklist_completed = sum(1 for item in checklist_items if item.completed)
    checklist_total = len(checklist_items)

    notifications = NotificationSetting.query.first()

    # Upcoming bills for this pay cycle. By default, bills due on payday
    # belong to the new cycle, but Settings can switch that behaviour.
    due_cutoff = cycle_bill_cutoff(settings, cycle_end, next_payday)
    due_before_next_payday = BillOccurrence.query.filter(
        BillOccurrence.due_date >= today.isoformat(),
        BillOccurrence.due_date <= due_cutoff.isoformat(),
        BillOccurrence.status == "Upcoming",
    ).order_by(BillOccurrence.due_date).all()

    due_next_30_days = BillOccurrence.query.filter(
        BillOccurrence.due_date >= today.isoformat(),
        BillOccurrence.due_date <= (today + timedelta(days=30)).isoformat(),
        BillOccurrence.status == "Upcoming",
    ).order_by(BillOccurrence.due_date).all()

    overdue = BillOccurrence.query.filter(
        BillOccurrence.due_date < today.isoformat(),
        BillOccurrence.status == "Upcoming",
    ).order_by(BillOccurrence.due_date).all()

    monthly_total = money(sum(annual_cost(b) for b in bills) / 12)
    annual_total = money(sum(annual_cost(b) for b in bills))

    latest_balance = AccountBalanceSnapshot.query.order_by(AccountBalanceSnapshot.snapshot_date.desc(), AccountBalanceSnapshot.id.desc()).first()
    current_cycle_unpaid = money(sum(o.amount for o in due_before_next_payday))
    projected_balance_after_cycle = money((latest_balance.balance if latest_balance else 0) - current_cycle_unpaid)

    setup_steps = [
        {"label": "Check household settings", "done": bool(settings.first_payday), "url": url_for("main.settings_page")},
        {"label": "Add your first recurring bill", "done": RecurringBill.query.count() > 0, "url": url_for("main.new_bill")},
        {"label": "Add an income source", "done": IncomeSource.query.count() > 0, "url": url_for("main.new_income")},
        {"label": "Review your buckets", "done": Bucket.query.count() > 0, "url": url_for("main.buckets")},
    ]
    show_setup = (not settings.setup_checklist_dismissed) and (not all(step["done"] for step in setup_steps))
    dashboard_widgets, enabled_widgets = get_dashboard_widgets()

    savings_goal_rows = []
    for p in purchases:
        pf = money(planned_purchase_fortnightly_amount(p, settings.first_payday))
        remaining = money(max(p.target_amount - p.amount_saved, 0))
        progress = min((p.amount_saved / p.target_amount * 100) if p.target_amount else 0, 100)
        fully_funded = p.amount_saved >= p.target_amount
        weeks_remaining = None
        if pf > 0 and not fully_funded:
            weeks_remaining = max(round((remaining / pf) * 2), 0)
        savings_goal_rows.append({
            "purchase": p,
            "per_fortnight": pf,
            "remaining": remaining,
            "progress": progress,
            "fully_funded": fully_funded,
            "weeks_remaining": weeks_remaining,
            "scope_label": planned_purchase_scope_label(p),
        })

    return render_template(
        "dashboard.html",
        bill_fortnightly_total=bill_fortnightly_total,
        purchase_fortnightly_total=purchase_fortnightly_total,
        buffer_amount=buffer_amount,
        total_set_aside=total_set_aside,
        cycle_start=cycle_start,
        cycle_end=cycle_end,
        next_payday=next_payday,
        due_cutoff=due_cutoff,
        due_before_next_payday=due_before_next_payday,
        due_next_30_days=due_next_30_days,
        overdue=overdue,
        monthly_total=monthly_total,
        annual_total=annual_total,
        purchases=purchases,
        shared_purchases=shared_purchases,
        individual_purchase_rows=individual_purchase_rows,
        planned_purchase_fortnightly_amount=planned_purchase_fortnightly_amount,
        setup_steps=setup_steps,
        show_setup=show_setup,
        latest_balance=latest_balance,
        current_cycle_unpaid=current_cycle_unpaid,
        projected_balance_after_cycle=projected_balance_after_cycle,
        income_sources=income_sources,
        income_items=income_items,
        income_total=income_total,
        buckets=buckets,
        bucket_allocations=bucket_allocations,
        total_bucket_amount=total_bucket_amount,
        remaining_after_buckets=remaining_after_buckets,
        bills_bucket_total=bills_bucket_total,
        bills_and_purchases_bucket_total=bills_and_purchases_bucket_total,
        bills_bucket_delta=bills_bucket_delta,
        bills_bucket_status=bills_bucket_status,
        checklist_items=checklist_items,
        checklist_completed=checklist_completed,
        checklist_total=checklist_total,
        notifications=notifications,
        person_bucket_allocations=person_bucket_allocations,
        dashboard_widgets=dashboard_widgets,
        enabled_widgets=enabled_widgets,
        savings_goal_rows=savings_goal_rows,
    )


@main.route("/dashboard/dismiss-setup", methods=["GET", "POST"])
@login_required
def dismiss_setup_checklist():
    settings = get_settings()
    settings.setup_checklist_dismissed = True
    db.session.commit()
    flash("Setup checklist dismissed.", "success")
    return redirect(url_for("main.dashboard"))




@main.route("/dashboard/layout", methods=["GET", "POST"])
@login_required
def dashboard_layout():
    widgets, enabled_widgets = get_dashboard_widgets()
    if request.method == "POST":
        for widget in widgets:
            widget.enabled = request.form.get(f"enabled_{widget.id}") == "on"
            try:
                widget.sort_order = int(request.form.get(f"sort_order_{widget.id}", widget.sort_order))
            except ValueError:
                pass
            size = request.form.get(f"size_{widget.id}", widget.size)
            if size in ["small", "medium", "wide"]:
                widget.size = size
        db.session.commit()
        flash("Dashboard layout updated.", "success")
        return redirect(url_for("main.dashboard_layout"))
    return render_template("dashboard_layout.html", widgets=widgets)


@main.route("/dashboard/layout/reset", methods=["POST"])
@login_required
def reset_dashboard_layout():
    DashboardWidget.query.delete()
    db.session.commit()
    defaults = [
        ("setup_checklist", "Setup checklist", True, 5, "wide", "Initial setup prompts. Can be hidden once the app is configured."),
        ("quick_links", "Quick links", True, 8, "small", "Shortcut buttons for Bills, Add Bill, Calendar, and Category Overview."),
        ("set_aside_summary", "Set-aside summary", True, 10, "wide", "Main fortnightly set-aside number and components."),
        ("income_summary", "Income summary", True, 20, "medium", "Expected household income and remaining amount after bucket transfers."),
        ("bucket_summary", "Bucket summary", True, 30, "medium", "Combined household bucket totals."),
        ("per_person_contributions", "Individual contributions", True, 40, "wide", "How each person contributes to the buckets this cycle."),
        ("bills_bucket_health", "Bills bucket health", True, 45, "medium", "Shows whether the bills bucket covers the fortnightly bills requirement."),
        ("payday_checklist", "Payday checklist", True, 48, "medium", "Quick link to the transfer checklist for payday."),
        ("due_before_next_payday", "Due this cycle", True, 50, "wide", "Upcoming bills due before the current cycle ends."),
        ("overdue_bills", "Overdue bills", True, 60, "wide", "Unpaid bills with due dates before today."),
        ("planned_purchases", "Planned purchases", False, 70, "medium", "Active planned purchases and quick-add saved amount."),
        ("account_balance", "Bills account balance", False, 80, "medium", "Latest manual bills account balance snapshot."),
        ("due_next_30_days", "Due in next 30 days", False, 90, "wide", "All unpaid bills due in the next 30 days."),
        ("recurring_totals", "Recurring totals", False, 100, "medium", "Monthly average and annual recurring bill totals."),
    ]
    for key, title, enabled, sort_order, size, description in defaults:
        db.session.add(DashboardWidget(widget_key=key, title=title, enabled=enabled, sort_order=sort_order, size=size, description=description))
    db.session.commit()
    flash("Dashboard layout reset.", "success")
    return redirect(url_for("main.dashboard_layout"))


@main.route("/settings", methods=["GET", "POST"])
@login_required
def settings_page():
    settings = get_settings()
    form = SettingsForm(obj=settings)
    if form.validate_on_submit():
        old_year = settings.budget_year
        form.populate_obj(settings)
        settings.first_payday = normalise_date_string(settings.first_payday)
        db.session.commit()
        # Regenerate all unpaid occurrences if the budget year changes.
        scope = "all_unpaid" if settings.budget_year != old_year else "future_unpaid"
        for bill in RecurringBill.query.all():
            regenerate_bill_occurrences(bill, scope=scope)
        db.session.commit()
        flash("Settings saved.", "success")
        return redirect(url_for("main.settings_page"))
    return render_template("settings.html", form=form)


@main.route("/bills")
@login_required
def bills():
    """Recurring bill list with column sorting and category filtering."""
    sort_key = request.args.get("sort", "name")
    sort_dir = request.args.get("dir", "asc")
    category_filter = request.args.get("category", "all")

    rows = RecurringBill.query.all()

    if category_filter not in [None, "", "all"]:
        if category_filter == "uncategorised":
            rows = [bill for bill in rows if not bill.category_id]
        else:
            try:
                category_id = int(category_filter)
                rows = [bill for bill in rows if bill.category_id == category_id]
            except ValueError:
                category_filter = "all"

    def bill_sort_value(bill):
        category_name = bill.category.name if bill.category else "Uncategorised"
        values = {
            "name": (bill.name or "").lower(),
            "category": category_name.lower(),
            "frequency": bill.frequency or "",
            "due": (bill.start_date or "", (bill.name or "").lower()),
            "amount": float(bill.amount or 0),
            "annual": annual_cost(bill),
            "fortnightly": fortnightly_bill_amount(bill),
            "active": 0 if bill.active else 1,
        }
        return values.get(sort_key, values["name"])

    reverse = sort_dir == "desc"
    rows = sorted(rows, key=bill_sort_value, reverse=reverse)

    categories = Category.query.filter(Category.active.is_(True)).order_by(Category.name).all()

    return render_template(
        "bills.html",
        bills=rows,
        categories=categories,
        selected_category=category_filter or "all",
        sort_key=sort_key,
        sort_dir=sort_dir,
        annual_cost=annual_cost,
        fortnightly_bill_amount=fortnightly_bill_amount,
    )


@main.route("/bills/category-overview")
@login_required
def bill_category_overview():
    """Show recurring bill spending totals grouped by category."""
    active_only = request.args.get("active", "1") != "0"
    include_set_aside_only = request.args.get("included", "0") == "1"

    rows = RecurringBill.query.all()
    if active_only:
        rows = [bill for bill in rows if bill.active]
    if include_set_aside_only:
        rows = [bill for bill in rows if bill.include_in_set_aside]

    category_budgets = {c.name: c.fortnightly_budget for c in Category.query.all()}

    grouped = {}
    for bill in rows:
        category_name = bill.category.name if bill.category else "Uncategorised"
        if category_name not in grouped:
            grouped[category_name] = {
                "category": category_name,
                "bill_count": 0,
                "weekly": 0,
                "fortnightly": 0,
                "monthly": 0,
                "yearly": 0,
                "budget": category_budgets.get(category_name),
            }
        yearly = annual_cost(bill)
        grouped[category_name]["bill_count"] += 1
        grouped[category_name]["weekly"] += yearly / 52
        grouped[category_name]["fortnightly"] += yearly / 26
        grouped[category_name]["monthly"] += yearly / 12
        grouped[category_name]["yearly"] += yearly

    summary_rows = []
    for row in grouped.values():
        row["weekly"] = money(row["weekly"])
        row["fortnightly"] = money(row["fortnightly"])
        row["monthly"] = money(row["monthly"])
        row["yearly"] = money(row["yearly"])
        if row["budget"] is not None:
            row["budget_delta"] = money(row["budget"] - row["fortnightly"])
        else:
            row["budget_delta"] = None
        summary_rows.append(row)

    summary_rows = sorted(summary_rows, key=lambda row: row["yearly"], reverse=True)

    totals = {
        "bill_count": sum(row["bill_count"] for row in summary_rows),
        "weekly": money(sum(row["weekly"] for row in summary_rows)),
        "fortnightly": money(sum(row["fortnightly"] for row in summary_rows)),
        "monthly": money(sum(row["monthly"] for row in summary_rows)),
        "yearly": money(sum(row["yearly"] for row in summary_rows)),
    }

    return render_template(
        "bill_category_overview.html",
        rows=summary_rows,
        totals=totals,
        active_only=active_only,
        include_set_aside_only=include_set_aside_only,
    )


@main.route("/bills/new", methods=["GET", "POST"])
@login_required
def new_bill():
    form = RecurringBillForm()
    form.category_id.choices = category_choices("Bill")
    if form.validate_on_submit():
        bill = RecurringBill()
        apply_bill_form(form, bill)
        new_category = get_or_create_category(form.new_category_name.data, "Bill")
        bill.category_id = new_category.id if new_category else (form.category_id.data or None)
        db.session.add(bill)
        db.session.flush()
        regenerate_bill_occurrences(bill, scope="all_unpaid")
        audit("add_bill", "RecurringBill", bill.name, f"Amount: {bill.amount}; frequency: {bill.frequency}")
        db.session.commit()
        flash("Bill added.", "success")
        return redirect(url_for("main.bills"))
    return render_template("bill_form.html", form=form, title="Add recurring bill", edit_mode=False)


@main.route("/bills/<int:bill_id>/edit", methods=["GET", "POST"])
@login_required
def edit_bill(bill_id):
    bill = db.get_or_404(RecurringBill, bill_id)
    form = RecurringBillForm(obj=bill)
    form.category_id.choices = category_choices("Bill")
    if request.method == "GET":
        form.category_id.data = bill.category_id or 0
        form.first_due_date.data = bill.start_date
    if form.validate_on_submit():
        scope = request.form.get("occurrence_update_scope", "future_unpaid")
        old_amount = bill.amount
        apply_bill_form(form, bill)
        new_category = get_or_create_category(form.new_category_name.data, "Bill")
        bill.category_id = new_category.id if new_category else (form.category_id.data or None)
        regenerate_bill_occurrences(bill, scope=scope)
        if old_amount is not None and abs((old_amount or 0) - (bill.amount or 0)) > 0.001:
            audit("amount_change", "RecurringBill", bill.name, f"Amount changed from {old_amount:.2f} to {bill.amount:.2f}")
        audit("edit_bill", "RecurringBill", bill.name, f"Occurrence scope: {scope}")
        db.session.commit()
        flash("Bill updated.", "success")
        return redirect(url_for("main.bills"))
    return render_template("bill_form.html", form=form, title="Edit recurring bill", edit_mode=True)


@main.route("/bills/<int:bill_id>/delete", methods=["POST"])
@login_required
def delete_bill(bill_id):
    bill = db.get_or_404(RecurringBill, bill_id)
    audit("delete_bill", "RecurringBill", bill.name if bill else "Unknown", "Deleted recurring bill")
    db.session.delete(bill)
    db.session.commit()
    flash("Bill deleted.", "success")
    return redirect(url_for("main.bills"))



@main.route("/bills/<int:bill_id>")
@login_required
def bill_detail(bill_id):
    bill = db.get_or_404(RecurringBill, bill_id)
    today_iso = date.today().isoformat()
    upcoming = BillOccurrence.query.filter(
        BillOccurrence.recurring_bill_id == bill.id,
        BillOccurrence.due_date >= today_iso,
    ).order_by(BillOccurrence.due_date).limit(12).all()
    history = BillOccurrence.query.filter(
        BillOccurrence.recurring_bill_id == bill.id,
        BillOccurrence.due_date < today_iso,
    ).order_by(BillOccurrence.due_date.desc()).limit(12).all()
    next_occurrence = upcoming[0] if upcoming else None
    amount_history = AuditLog.query.filter(
        AuditLog.action == "amount_change",
        AuditLog.entity_type == "RecurringBill",
        AuditLog.entity_name == bill.name,
    ).order_by(AuditLog.created_at.desc()).limit(20).all()
    return render_template(
        "bill_detail.html",
        bill=bill,
        upcoming=upcoming,
        history=history,
        next_occurrence=next_occurrence,
        annual=annual_cost(bill),
        per_fortnight=fortnightly_bill_amount(bill),
        amount_history=amount_history,
    )


@main.route("/cycle-closeout", methods=["GET", "POST"])
@login_required
def cycle_closeout():
    settings = get_settings()
    cycle_choice = request.args.get("cycle", "current")
    cycle_offset = 1 if cycle_choice == "next" else 0
    income_sources = IncomeSource.query.filter_by(active=True).order_by(IncomeSource.next_pay_date, IncomeSource.name).all()
    cycle_start, cycle_end, next_payday = get_cycle_window(settings, income_sources, offset=cycle_offset)
    closeout = get_or_create_cycle_closeout(cycle_start, cycle_end)
    form = CycleCloseoutForm(obj=closeout)

    occurrences = get_cycle_occurrences(settings, cycle_start, cycle_end, next_payday)
    paid = [o for o in occurrences if o.status == "Paid"]
    skipped = [o for o in occurrences if o.status == "Skipped"]
    unpaid = [o for o in occurrences if o.status == "Upcoming"]
    income_items, income_total = income_for_cycle(income_sources, cycle_start, cycle_end)
    individual_income_total = money(sum(
        item["amount"] for item in income_items
        if getattr(item["source"], "income_scope", "Individual") != "Shared"
    ))
    buckets = Bucket.query.filter_by(active=True).order_by(Bucket.sort_order, Bucket.name).all()
    bucket_allocations = calculate_bucket_allocations(buckets, individual_income_total)
    person_bucket_allocations = calculate_person_bucket_allocations(income_items, buckets, individual_income_total)
    shared_bucket_additions, shared_standard_pool = calculate_shared_income_bucket_additions(income_items, buckets)
    if shared_standard_pool:
        standard_shared_allocations = calculate_bucket_allocations(buckets, shared_standard_pool)
        for i, row in enumerate(bucket_allocations):
            row["rounded_amount"] = money(row["rounded_amount"] + standard_shared_allocations[i]["rounded_amount"])
            row["raw_amount"] = money(row["raw_amount"] + standard_shared_allocations[i]["raw_amount"])
    for row in bucket_allocations:
        extra = shared_bucket_additions.get(row["bucket"].id, 0)
        if extra:
            row["rounded_amount"] = money(row["rounded_amount"] + extra)
            row["raw_amount"] = money(row["raw_amount"] + extra)
    checklist_items = ensure_payday_checklist_items(settings, cycle_start, income_items, person_bucket_allocations, bucket_allocations)

    if form.validate_on_submit():
        closeout.notes = form.notes.data
        closeout.actual_income = form.actual_income.data
        closeout.status = "Closed"
        closeout.closed_at = datetime.now().isoformat(timespec="seconds")
        audit("close_cycle", "CycleCloseout", cycle_start.isoformat(), f"Closed pay cycle {cycle_start.isoformat()} to {cycle_end.isoformat()}")
        db.session.commit()
        flash("Pay cycle closed.", "success")
        return redirect(url_for("main.cycle_closeout", cycle=cycle_choice))

    return render_template(
        "cycle_closeout.html",
        form=form,
        cycle_choice=cycle_choice,
        cycle_start=cycle_start,
        cycle_end=cycle_end,
        next_payday=next_payday,
        closeout=closeout,
        occurrences=occurrences,
        paid=paid,
        skipped=skipped,
        unpaid=unpaid,
        income_items=income_items,
        income_total=income_total,
        bucket_allocations=bucket_allocations,
        person_bucket_allocations=person_bucket_allocations,
        individual_purchase_rows=individual_purchase_contributions(settings.first_payday),
        checklist_items=checklist_items,
    )


@main.route("/system-info")
@login_required
def system_info():
    settings = get_settings()
    income_sources = IncomeSource.query.filter_by(active=True).all()
    cycle_start, cycle_end, next_payday = get_cycle_window(settings, income_sources)
    db_path = get_database_path()

    database_rows = [
        {"label": "Database URI", "value": current_app.config.get("SQLALCHEMY_DATABASE_URI")},
        {"label": "Database path", "value": db_path},
        {"label": "Database exists", "value": "Yes" if os.path.exists(db_path) else "No"},
        {"label": "Database size", "value": file_size_label(db_path)},
    ]

    app_rows = [
        {"label": "App version", "value": APP_VERSION},
        {"label": "Release", "value": APP_RELEASE_NAME},
        {"label": "Git commit", "value": get_git_commit()},
        {"label": "Python", "value": platform.python_version()},
        {"label": "Flask debug", "value": "On" if current_app.debug else "Off"},
    ]

    data_rows = [
        {"label": "Active bills", "value": RecurringBill.query.filter_by(active=True).count()},
        {"label": "Active income sources", "value": IncomeSource.query.filter_by(active=True).count()},
        {"label": "Active buckets", "value": Bucket.query.filter_by(active=True).count()},
        {"label": "Active planned purchases", "value": PlannedPurchase.query.filter_by(status="Active").count()},
        {"label": "Unpaid bill occurrences", "value": BillOccurrence.query.filter(BillOccurrence.status == "Upcoming").count()},
        {"label": "Overdue unpaid occurrences", "value": BillOccurrence.query.filter(BillOccurrence.status == "Upcoming", BillOccurrence.due_date < date.today().isoformat()).count()},
    ]

    cycle_rows = [
        {"label": "Current cycle", "value": f"{cycle_start.strftime('%d %b %Y')} to {cycle_end.strftime('%d %b %Y')}"},
        {"label": "Next payday", "value": next_payday.strftime('%d %b %Y')},
        {"label": "Payday bill handling", "value": getattr(settings, "payday_bill_handling", "new_cycle")},
    ]

    return render_template(
        "system_info.html",
        app_rows=app_rows,
        database_rows=database_rows,
        data_rows=data_rows,
        cycle_rows=cycle_rows,
    )


@main.route("/health-check")
@login_required
def health_check_page():
    settings = get_settings()
    checks = []
    active_bills = RecurringBill.query.filter_by(active=True).all()
    active_income = IncomeSource.query.filter_by(active=True).all()
    active_buckets = Bucket.query.filter_by(active=True).order_by(Bucket.sort_order, Bucket.name).all()
    overdue = BillOccurrence.query.filter(BillOccurrence.due_date < date.today().isoformat(), BillOccurrence.status == "Upcoming").count()

    checks.append({"status": "ok" if active_income else "warning", "title": "Income sources", "detail": "At least one active income source is configured." if active_income else "No active income sources are configured."})
    checks.append({"status": "ok" if active_bills else "warning", "title": "Recurring bills", "detail": f"{len(active_bills)} active recurring bills found." if active_bills else "No active recurring bills are configured."})
    uncategorised = [b for b in active_bills if not b.category_id]
    checks.append({"status": "ok" if not uncategorised else "warning", "title": "Bill categories", "detail": "All active bills have categories." if not uncategorised else f"{len(uncategorised)} active bills are uncategorised."})
    capped = [b for b in active_buckets if b.cap_to_remaining]
    checks.append({"status": "ok" if len(capped) <= 1 else "danger", "title": "Remainder bucket", "detail": "Remainder bucket rule is valid." if len(capped) <= 1 else "More than one active bucket uses the remainder option."})
    percentage_total = money(sum(b.percentage for b in active_buckets if b.fixed_amount in [None, ""]))
    checks.append({"status": "ok" if 95 <= percentage_total <= 105 else "warning", "title": "Percentage buckets", "detail": f"Active percentage bucket total is {percentage_total:.2f}%."})
    checks.append({"status": "ok" if overdue == 0 else "warning", "title": "Overdue bills", "detail": "No overdue unpaid bills." if overdue == 0 else f"{overdue} bills are overdue and unpaid."})
    checks.append({"status": "ok" if getattr(settings, "payday_bill_handling", "new_cycle") in ["new_cycle", "previous_cycle"] else "warning", "title": "Payday bill setting", "detail": "Bills due on payday handling is configured."})

    db_path = get_database_path()
    checks.append({"status": "ok" if os.path.exists(db_path) else "danger", "title": "Database file", "detail": f"Database path: {db_path}"})
    checks.append({"status": "ok", "title": "App version", "detail": f"{APP_VERSION} — {APP_RELEASE_NAME}"})

    return render_template("health_check.html", checks=checks)


@main.route("/purchases")
@login_required
def purchases():
    priority_order = {"High": 0, "Medium": 1, "Low": 2}
    rows = PlannedPurchase.query.all()
    rows = sorted(rows, key=lambda p: (p.status != "Active", priority_order.get(p.priority, 1), p.target_date or "9999-12-31", p.name.lower()))
    settings = get_settings()

    purchase_rows = planned_purchase_summary_rows(rows, settings.first_payday)
    shared_rows = [row for row in purchase_rows if row["is_shared"]]
    individual_rows = [row for row in purchase_rows if not row["is_shared"]]

    individual_purchase_rows = individual_purchase_contributions(settings.first_payday)
    individual_people = []
    for person_row in individual_purchase_rows:
        person_rows = [row for row in individual_rows if row["owner"] == person_row["person"]]
        individual_people.append({
            "person": person_row["person"],
            "purchases": person_rows,
            "total": person_row["total"],
            "totals": planned_purchase_totals(person_rows),
        })

    unassigned_individual_rows = [row for row in individual_rows if row["owner"] == "Unassigned"]
    if unassigned_individual_rows:
        individual_people.append({
            "person": "Unassigned",
            "purchases": unassigned_individual_rows,
            "total": money(sum(row["per_fortnight"] for row in unassigned_individual_rows if row["purchase"].status == "Active")),
            "totals": planned_purchase_totals(unassigned_individual_rows),
        })

    return render_template(
        "purchases.html",
        purchases=rows,
        purchase_rows=purchase_rows,
        shared_rows=shared_rows,
        individual_rows=individual_rows,
        shared_totals=planned_purchase_totals(shared_rows),
        all_totals=planned_purchase_totals(purchase_rows),
        individual_people=individual_people,
        planned_purchase_fortnightly_amount=planned_purchase_fortnightly_amount,
        first_payday=settings.first_payday,
        individual_purchase_rows=individual_purchase_rows,
    )


@main.route("/purchases/new", methods=["GET", "POST"])
@login_required
def new_purchase():
    form = PlannedPurchaseForm()
    prepare_purchase_form(form)
    if form.validate_on_submit():
        purchase = PlannedPurchase()
        form.populate_obj(purchase)
        if purchase.purchase_scope != "Individual":
            purchase.owner_name = None
        purchase.target_date = normalise_date_string(purchase.target_date)
        new_category = get_or_create_category(form.new_category_name.data, "Purchase")
        purchase.category_id = new_category.id if new_category else (form.category_id.data or None)
        db.session.add(purchase)
        db.session.commit()
        flash("Planned purchase added.", "success")
        return redirect(url_for("main.purchases"))
    return render_template("purchase_form.html", form=form, title="Add planned purchase")


@main.route("/purchases/<int:purchase_id>/edit", methods=["GET", "POST"])
@login_required
def edit_purchase(purchase_id):
    purchase = db.get_or_404(PlannedPurchase, purchase_id)
    form = PlannedPurchaseForm(obj=purchase)
    prepare_purchase_form(form, purchase)
    if form.validate_on_submit():
        form.populate_obj(purchase)
        if purchase.purchase_scope != "Individual":
            purchase.owner_name = None
        purchase.target_date = normalise_date_string(purchase.target_date)
        new_category = get_or_create_category(form.new_category_name.data, "Purchase")
        purchase.category_id = new_category.id if new_category else (form.category_id.data or None)
        db.session.commit()
        flash("Planned purchase updated.", "success")
        return redirect(url_for("main.purchases"))
    return render_template("purchase_form.html", form=form, title="Edit planned purchase")


@main.route("/purchases/<int:purchase_id>/delete", methods=["POST"])
@login_required
def delete_purchase(purchase_id):
    purchase = db.get_or_404(PlannedPurchase, purchase_id)
    db.session.delete(purchase)
    db.session.commit()
    flash("Planned purchase deleted.", "success")
    return redirect(url_for("main.purchases"))


@main.route("/purchases/<int:purchase_id>/add-saved", methods=["POST"])
@login_required
def add_purchase_saved(purchase_id):
    purchase = db.get_or_404(PlannedPurchase, purchase_id)
    try:
        amount = float(request.form.get("amount", 0))
    except ValueError:
        amount = 0

    if amount <= 0:
        flash("Enter an amount greater than zero.", "warning")
        return redirect(request.referrer or url_for("main.purchases"))

    purchase.amount_saved = money(min(purchase.amount_saved + amount, purchase.target_amount))
    if purchase.amount_saved >= purchase.target_amount:
        flash("Saved amount updated. This target is now fully funded.", "success")
    else:
        flash("Saved amount updated.", "success")
    db.session.commit()
    return redirect(request.referrer or url_for("main.purchases"))


@main.route("/purchases/<int:purchase_id>/mark-purchased", methods=["POST"])
@login_required
def mark_purchase_purchased(purchase_id):
    purchase = db.get_or_404(PlannedPurchase, purchase_id)
    purchase.status = "Purchased"
    purchase.amount_saved = max(purchase.amount_saved, purchase.target_amount)
    db.session.commit()
    flash("Planned purchase marked as purchased.", "success")
    return redirect(request.referrer or url_for("main.purchases"))


@main.route("/calendar")
@login_required
def calendar_current():
    settings = get_settings()
    today = date.today()
    year = settings.budget_year or today.year
    month = today.month if year == today.year else 1
    return redirect(url_for("main.month_view", year=year, month=month, view="calendar"))


@main.route("/month/<int:year>/<int:month>")
@login_required
def month_view(year, month):
    start = date(year, month, 1)
    if month == 12:
        end = date(year, 12, 31)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)

    selected_view = request.args.get("view", "calendar")
    if selected_view not in ["calendar", "list"]:
        selected_view = "calendar"

    occurrences = BillOccurrence.query.filter(
        BillOccurrence.due_date >= start.isoformat(),
        BillOccurrence.due_date <= end.isoformat(),
    ).order_by(BillOccurrence.due_date).all()

    income_sources = IncomeSource.query.filter_by(active=True).order_by(IncomeSource.owner_name, IncomeSource.name).all()
    income_events = []
    for source in income_sources:
        for pay_date in generate_income_dates(source, start, end):
            income_events.append({
                "date": pay_date,
                "type": "income",
                "title": f"{source.owner_name} - {source.name}",
                "amount": money(source.amount),
                "status": "Expected",
                "source": source,
            })

    bill_events = []
    for occurrence in occurrences:
        bill_events.append({
            "date": parse_date(occurrence.due_date),
            "type": "bill",
            "title": occurrence.bill.name,
            "amount": money(occurrence.amount),
            "status": occurrence.status,
            "occurrence": occurrence,
            "category": occurrence.bill.category.name if occurrence.bill.category else "Uncategorised",
        })

    month_events = sorted(bill_events + income_events, key=lambda item: (item["date"], 0 if item["type"] == "income" else 1, item["title"]))
    events_by_date = {}
    for event in month_events:
        events_by_date.setdefault(event["date"].isoformat(), []).append(event)

    today = date.today()
    calendar_weeks = []
    for week in Calendar(firstweekday=0).monthdatescalendar(year, month):
        calendar_weeks.append([
            {
                "date": day,
                "in_month": day.month == month,
                "is_today": day == today,
                "events": events_by_date.get(day.isoformat(), []),
            }
            for day in week
        ])

    total = money(sum(o.amount for o in occurrences))
    paid = money(sum(o.amount for o in occurrences if o.status == "Paid"))
    unpaid = money(total - paid)
    income_total = money(sum(event["amount"] for event in income_events))

    overdue_global = BillOccurrence.query.filter(
        BillOccurrence.due_date < today.isoformat(),
        BillOccurrence.status == "Upcoming",
    ).order_by(BillOccurrence.due_date).all()

    return render_template(
        "month.html",
        year=year,
        month=month,
        month_label=month_name[month],
        month_names=list(month_name),
        selected_view=selected_view,
        occurrences=occurrences,
        month_events=month_events,
        calendar_weeks=calendar_weeks,
        total=total,
        paid=paid,
        unpaid=unpaid,
        income_total=income_total,
        today=today,
        overdue_global=overdue_global,
    )


@main.route("/occurrences/<int:occurrence_id>/paid", methods=["POST"])
@login_required
def mark_occurrence_paid(occurrence_id):
    occurrence = db.get_or_404(BillOccurrence, occurrence_id)
    occurrence.status = "Paid"
    occurrence.paid_date = date.today().isoformat()
    audit("mark_bill_paid", "BillOccurrence", occurrence.bill.name if occurrence.bill else "Bill", occurrence.due_date)
    db.session.commit()
    flash("Bill marked as paid.", "success")
    return redirect(request.form.get("return_to") or request.referrer or url_for("main.dashboard"))


@main.route("/occurrences/<int:occurrence_id>/unpaid", methods=["POST"])
@login_required
def mark_occurrence_unpaid(occurrence_id):
    occurrence = db.get_or_404(BillOccurrence, occurrence_id)
    occurrence.status = "Upcoming"
    occurrence.paid_date = None
    audit("mark_bill_unpaid", "BillOccurrence", occurrence.bill.name if occurrence.bill else "Bill", occurrence.due_date)
    db.session.commit()
    flash("Bill marked as unpaid.", "success")
    return redirect(request.form.get("return_to") or request.referrer or url_for("main.dashboard"))


@main.route("/occurrences/<int:occurrence_id>/skip", methods=["POST"])
@login_required
def skip_occurrence(occurrence_id):
    occurrence = db.get_or_404(BillOccurrence, occurrence_id)
    occurrence.status = "Skipped"
    occurrence.paid_date = None
    audit("skip_bill", "BillOccurrence", occurrence.bill.name if occurrence.bill else "Bill", occurrence.due_date)
    db.session.commit()
    flash("Bill occurrence skipped.", "success")
    return redirect(request.form.get("return_to") or request.referrer or url_for("main.dashboard"))


@main.route("/pay-cycle")
@login_required
def pay_cycle():
    settings = get_settings()
    cycle_choice = request.args.get("cycle", "current")
    cycle_offset = 1 if cycle_choice == "next" else 0
    income_sources = IncomeSource.query.filter_by(active=True).order_by(IncomeSource.next_pay_date, IncomeSource.name).all()
    cycle_start, cycle_end, next_payday = get_cycle_window(settings, income_sources, offset=cycle_offset)
    occurrences = get_cycle_occurrences(settings, cycle_start, cycle_end, next_payday)

    bills_due = money(sum(o.amount for o in occurrences if o.status != "Paid"))
    active_bills = RecurringBill.query.filter_by(active=True).all()
    purchases = shared_active_purchases()
    recurring_average = money(sum(fortnightly_bill_amount(b) for b in active_bills))
    purchase_average = money(sum(planned_purchase_fortnightly_amount(p, settings.first_payday) for p in purchases))
    total_average = money(recurring_average + purchase_average + settings.default_buffer_amount)
    income_items, income_total = income_for_cycle(income_sources, cycle_start, cycle_end)
    buckets = Bucket.query.filter_by(active=True).order_by(Bucket.sort_order, Bucket.name).all()
    bucket_allocations = calculate_bucket_allocations(buckets, income_total)
    person_bucket_allocations = calculate_person_bucket_allocations(income_items, buckets, income_total)
    closeout = CycleCloseout.query.filter_by(cycle_start=cycle_start.isoformat()).first()

    return render_template(
        "pay_cycle.html",
        cycle_choice=cycle_choice,
        cycle_start=cycle_start,
        cycle_end=cycle_end,
        next_payday=next_payday,
        due_cutoff=cycle_bill_cutoff(settings, cycle_end, next_payday),
        closeout=closeout,
        occurrences=occurrences,
        bills_due=bills_due,
        recurring_average=recurring_average,
        purchase_average=purchase_average,
        total_average=total_average,
        income_items=income_items,
        income_total=income_total,
        bucket_allocations=bucket_allocations,
        person_bucket_allocations=person_bucket_allocations,
    )


@main.route("/income")
@login_required
def income_sources():
    rows = IncomeSource.query.order_by(IncomeSource.active.desc(), IncomeSource.owner_name, IncomeSource.next_pay_date, IncomeSource.name).all()
    today = date.today()
    income_rows = []
    for income in rows:
        upcoming = next_income_pay_date(income, today=today) if income.active else None
        income_rows.append({
            "income": income,
            "upcoming_pay_date": upcoming,
            "anchor_pay_date": parse_date(income.next_pay_date) if income.next_pay_date else None,
        })
    return render_template("income.html", income_rows=income_rows)


def _bucket_choices():
    """Return (id, name) choices for all active buckets, for use in forms."""
    return [(b.id, b.name) for b in Bucket.query.filter_by(active=True).order_by(Bucket.sort_order, Bucket.name).all()]


@main.route("/income/new", methods=["GET", "POST"])
@login_required
def new_income():
    form = IncomeSourceForm()
    form.lump_bucket_id.choices = [(0, "— select bucket —")] + _bucket_choices()
    if form.validate_on_submit():
        income = IncomeSource()
        form.populate_obj(income)
        income.next_pay_date = normalise_date_string(income.next_pay_date)
        # Clear fields that don't apply to individual income.
        if income.income_scope == "Individual":
            income.allocation_mode = "standard"
            income.lump_bucket_id = None
        else:
            income.owner_name = "Household"
            if income.lump_bucket_id == 0:
                income.lump_bucket_id = None
        db.session.add(income)
        db.session.commit()
        audit("add_income", "IncomeSource", income.name, f"Scope: {income.income_scope}, Freq: {income.frequency}")
        flash("Income source added.", "success")
        return redirect(url_for("main.income_sources"))
    return render_template("income_form.html", form=form, title="Add income source")


@main.route("/income/<int:income_id>/edit", methods=["GET", "POST"])
@login_required
def edit_income(income_id):
    income = db.get_or_404(IncomeSource, income_id)
    form = IncomeSourceForm(obj=income)
    form.lump_bucket_id.choices = [(0, "— select bucket —")] + _bucket_choices()
    if form.validate_on_submit():
        form.populate_obj(income)
        income.next_pay_date = normalise_date_string(income.next_pay_date)
        if income.income_scope == "Individual":
            income.allocation_mode = "standard"
            income.lump_bucket_id = None
        else:
            income.owner_name = "Household"
            if income.lump_bucket_id == 0:
                income.lump_bucket_id = None
        db.session.commit()
        audit("edit_income", "IncomeSource", income.name, f"Scope: {income.income_scope}, Freq: {income.frequency}")
        flash("Income source updated.", "success")
        return redirect(url_for("main.income_sources"))
    allocations = SharedIncomeAllocation.query.filter_by(income_source_id=income_id).order_by(SharedIncomeAllocation.sort_order).all()
    alloc_form = SharedIncomeAllocationForm()
    alloc_form.bucket_id.choices = _bucket_choices()
    return render_template("income_form.html", form=form, title="Edit income source", income=income, allocations=allocations, alloc_form=alloc_form)


@main.route("/income/<int:income_id>/delete", methods=["POST"])
@login_required
def delete_income(income_id):
    income = db.get_or_404(IncomeSource, income_id)
    db.session.delete(income)
    db.session.commit()
    audit("delete_income", "IncomeSource", income.name, None)
    flash("Income source deleted.", "success")
    return redirect(url_for("main.income_sources"))


@main.route("/income/<int:income_id>/allocations/add", methods=["POST"])
@login_required
def add_income_allocation(income_id):
    """Add a custom bucket allocation row to a shared income source."""
    income = db.get_or_404(IncomeSource, income_id)
    if income.income_scope != "Shared" or income.allocation_mode != "custom":
        flash("Custom allocations only apply to shared income sources in custom mode.", "warning")
        return redirect(url_for("main.edit_income", income_id=income_id))

    form = SharedIncomeAllocationForm()
    form.bucket_id.choices = _bucket_choices()
    if form.validate_on_submit():
        # Enforce single remainder bucket per income source.
        if form.is_remainder.data:
            SharedIncomeAllocation.query.filter_by(
                income_source_id=income_id, is_remainder=True
            ).update({"is_remainder": False})
        alloc = SharedIncomeAllocation(income_source_id=income_id)
        form.populate_obj(alloc)
        db.session.add(alloc)
        db.session.commit()
        flash("Allocation added.", "success")
    else:
        flash("Could not add allocation — check the form.", "warning")
    return redirect(url_for("main.edit_income", income_id=income_id))


@main.route("/income/<int:income_id>/allocations/<int:alloc_id>/delete", methods=["POST"])
@login_required
def delete_income_allocation(income_id, alloc_id):
    """Remove a custom bucket allocation row."""
    alloc = db.get_or_404(SharedIncomeAllocation, alloc_id)
    if alloc.income_source_id != income_id:
        flash("Allocation does not belong to this income source.", "danger")
        return redirect(url_for("main.edit_income", income_id=income_id))
    db.session.delete(alloc)
    db.session.commit()
    flash("Allocation removed.", "success")
    return redirect(url_for("main.edit_income", income_id=income_id))


@main.route("/buckets", methods=["GET", "POST"])
@login_required
def buckets():
    enforce_single_remainder_bucket()
    form = BucketForm()
    if form.validate_on_submit():
        bucket = Bucket()
        if apply_bucket_form(bucket, form):
            db.session.add(bucket)
            db.session.flush()
            if bucket.cap_to_remaining:
                enforce_single_remainder_bucket(preferred_bucket=bucket)
            db.session.commit()
            flash("Bucket added.", "success")
            return redirect(url_for("main.buckets"))
    elif request.method == "GET":
        form.allocation_method.data = "percentage"

    rows = Bucket.query.order_by(Bucket.active.desc(), Bucket.sort_order, Bucket.name).all()
    active_buckets = [bucket for bucket in rows if bucket.active]
    total_percentage = money(sum(b.percentage for b in rows if b.active and b.fixed_amount in [None, ""]))

    settings = get_settings()
    income_sources = IncomeSource.query.filter_by(active=True).order_by(IncomeSource.next_pay_date, IncomeSource.owner_name, IncomeSource.name).all()
    cycle_start, cycle_end, next_payday = get_cycle_window(settings, income_sources)
    income_items, income_total = income_for_cycle(income_sources, cycle_start, cycle_end)
    bucket_allocations = calculate_bucket_allocations(active_buckets, income_total)
    person_bucket_allocations = calculate_person_bucket_allocations(income_items, active_buckets, income_total)

    person_names = [person["person"] for person in person_bucket_allocations]
    bucket_contribution_rows = []
    for index, combined in enumerate(bucket_allocations):
        people = []
        for person in person_bucket_allocations:
            allocation = person["bucket_allocations"][index]
            people.append({
                "person": person["person"],
                "amount": allocation["rounded_amount"],
                "capped": allocation["capped"],
            })
        bucket_contribution_rows.append({
            "bucket": combined["bucket"],
            "target_label": combined["target_label"],
            "combined_amount": combined["rounded_amount"],
            "percentage_of_income": combined["percentage_of_income"],
            "capped": combined["capped"],
            "people": people,
        })

    bucket_total = money(sum(row["rounded_amount"] for row in bucket_allocations))
    remaining_after_buckets = money(income_total - bucket_total)

    return render_template(
        "buckets.html",
        form=form,
        buckets=rows,
        total_percentage=total_percentage,
        cycle_start=cycle_start,
        cycle_end=cycle_end,
        next_payday=next_payday,
        income_total=income_total,
        bucket_total=bucket_total,
        remaining_after_buckets=remaining_after_buckets,
        bucket_contribution_rows=bucket_contribution_rows,
        person_bucket_allocations=person_bucket_allocations,
        person_names=person_names,
    )


@main.route("/buckets/<int:bucket_id>/edit", methods=["GET", "POST"])
@login_required
def edit_bucket(bucket_id):
    bucket = db.get_or_404(Bucket, bucket_id)
    form = BucketForm(obj=bucket)
    if form.validate_on_submit():
        if apply_bucket_form(bucket, form):
            if bucket.cap_to_remaining:
                enforce_single_remainder_bucket(preferred_bucket=bucket)
            db.session.commit()
            flash("Bucket updated.", "success")
            return redirect(url_for("main.buckets"))
    elif request.method == "GET":
        prepare_bucket_form_for_display(form, bucket)
    return render_template("bucket_form.html", form=form, title="Edit bucket")


@main.route("/buckets/<int:bucket_id>/delete", methods=["POST"])
@login_required
def delete_bucket(bucket_id):
    bucket = db.get_or_404(Bucket, bucket_id)
    db.session.delete(bucket)
    db.session.commit()
    flash("Bucket deleted.", "success")
    return redirect(url_for("main.buckets"))


@main.route("/pay-split")
@login_required
def pay_split():
    settings = get_settings()
    income_sources = IncomeSource.query.filter_by(active=True).order_by(IncomeSource.next_pay_date, IncomeSource.name).all()
    cycle_start, cycle_end, next_payday = get_cycle_window(settings, income_sources)
    income_items, income_total = income_for_cycle(income_sources, cycle_start, cycle_end)
    buckets = Bucket.query.filter_by(active=True).order_by(Bucket.sort_order, Bucket.name).all()
    bucket_allocations = calculate_bucket_allocations(buckets, income_total)
    person_bucket_allocations = calculate_person_bucket_allocations(income_items, buckets, income_total)
    recurring_average = money(sum(fortnightly_bill_amount(b) for b in RecurringBill.query.filter_by(active=True).all()))
    purchase_average = money(sum(planned_purchase_fortnightly_amount(p, settings.first_payday) for p in shared_active_purchases()))
    required_set_aside = money(recurring_average + purchase_average + settings.default_buffer_amount)
    bills_bucket_total = money(sum(row["rounded_amount"] for row in bucket_allocations if row["bucket"].bucket_type in ["Bills", "Planned purchases"]))
    bucket_total = money(sum(row["rounded_amount"] for row in bucket_allocations))
    remaining = money(income_total - bucket_total)
    shortfall = money(required_set_aside - bills_bucket_total)
    return render_template(
        "pay_split.html",
        cycle_start=cycle_start,
        cycle_end=cycle_end,
        next_payday=next_payday,
        income_items=income_items,
        income_total=income_total,
        bucket_allocations=bucket_allocations,
        person_bucket_allocations=person_bucket_allocations,
        individual_purchase_rows=individual_purchase_contributions(settings.first_payday),
        recurring_average=recurring_average,
        purchase_average=purchase_average,
        required_set_aside=required_set_aside,
        bills_bucket_total=bills_bucket_total,
        shortfall=shortfall,
        bucket_total=bucket_total,
        remaining=remaining,
    )


@main.route("/categories", methods=["GET", "POST"])
@login_required
def categories():
    form = CategoryForm()
    if form.validate_on_submit():
        category = Category()
        form.populate_obj(category)
        db.session.add(category)
        db.session.commit()
        flash("Category added.", "success")
        return redirect(url_for("main.categories"))
    rows = Category.query.order_by(Category.name).all()
    return render_template("categories.html", form=form, categories=rows)


@main.route("/account-balance", methods=["GET", "POST"])
@login_required
def account_balance():
    form = AccountBalanceForm(snapshot_date=date.today().isoformat())
    if form.validate_on_submit():
        snapshot = AccountBalanceSnapshot(
            snapshot_date=normalise_date_string(form.snapshot_date.data),
            balance=money(form.balance.data),
            notes=form.notes.data,
        )
        db.session.add(snapshot)
        db.session.commit()
        flash("Account balance saved.", "success")
        return redirect(url_for("main.account_balance"))
    rows = AccountBalanceSnapshot.query.order_by(AccountBalanceSnapshot.snapshot_date.desc(), AccountBalanceSnapshot.id.desc()).limit(20).all()
    return render_template("account_balance.html", form=form, snapshots=rows)



@main.route("/payday-checklist", methods=["GET", "POST"])
@login_required
def payday_checklist():
    settings = get_settings()
    cycle_choice = request.args.get("cycle", "current")
    cycle_offset = 1 if cycle_choice == "next" else 0
    income_sources = IncomeSource.query.filter_by(active=True).order_by(IncomeSource.next_pay_date, IncomeSource.name).all()
    cycle_start, cycle_end, next_payday = get_cycle_window(settings, income_sources, offset=cycle_offset)
    income_items, income_total = income_for_cycle(income_sources, cycle_start, cycle_end)
    buckets = Bucket.query.filter_by(active=True).order_by(Bucket.sort_order, Bucket.name).all()
    bucket_allocations = calculate_bucket_allocations(buckets, income_total)
    person_bucket_allocations = calculate_person_bucket_allocations(income_items, buckets, income_total)
    checklist_items = ensure_payday_checklist_items(settings, cycle_start, income_items, person_bucket_allocations, bucket_allocations)
    hidden_preferences = PaydayChecklistPreference.query.filter_by(hidden=True).order_by(PaydayChecklistPreference.label).all()
    cycle_occurrences = get_cycle_occurrences(settings, cycle_start, cycle_end, next_payday)
    cycle_unpaid = [o for o in cycle_occurrences if o.status == "Upcoming"]
    cycle_unpaid_total = money(sum(o.amount for o in cycle_unpaid))
    checklist_rows = [
        {
            "item": item,
            "is_transfer": bool(item.item_key and item.item_key.startswith("transfer_")),
        }
        for item in checklist_items
    ]

    if request.method == "POST":
        for item in checklist_items:
            completed = request.form.get(f"completed_{item.id}") == "on"
            if completed and not item.completed:
                item.completed_at = datetime.now().isoformat(timespec="seconds")
            elif not completed:
                item.completed_at = None
            item.completed = completed
        audit("update_payday_checklist", "PaydayChecklist", cycle_start.isoformat(), "Updated payday checklist")
        db.session.commit()
        flash("Payday checklist updated.", "success")
        return redirect(url_for("main.payday_checklist", cycle=cycle_choice))

    return render_template(
        "payday_checklist.html",
        cycle_choice=cycle_choice,
        cycle_start=cycle_start,
        cycle_end=cycle_end,
        next_payday=next_payday,
        due_cutoff=cycle_bill_cutoff(settings, cycle_end, next_payday),
        income_items=income_items,
        income_total=income_total,
        bucket_allocations=bucket_allocations,
        person_bucket_allocations=person_bucket_allocations,
        individual_purchase_rows=individual_purchase_contributions(settings.first_payday),
        checklist_items=checklist_items,
        checklist_rows=checklist_rows,
        hidden_preferences=hidden_preferences,
        cycle_occurrences=cycle_occurrences,
        cycle_unpaid=cycle_unpaid,
        cycle_unpaid_total=cycle_unpaid_total,
    )


@main.route("/payday-checklist/items/<int:item_id>/hide", methods=["POST"])
@login_required
def hide_payday_checklist_item(item_id):
    item = db.get_or_404(PaydayChecklistItem, item_id)
    if not item:
        flash("Checklist item not found.", "warning")
        return redirect(url_for("main.payday_checklist"))
    pref = PaydayChecklistPreference.query.filter_by(item_key=item.item_key).first()
    if not pref:
        pref = PaydayChecklistPreference(item_key=item.item_key, label=item.label, hidden=True, reason="automatic_transfer")
        db.session.add(pref)
    else:
        pref.label = item.label
        pref.hidden = True
        pref.reason = "automatic_transfer"
    audit("hide_payday_checklist_item", "PaydayChecklistPreference", item.label, "Hidden from future payday checklists")
    db.session.commit()
    flash("Checklist item hidden. You can restore it from the hidden automatic transfers section.", "success")
    return redirect(url_for("main.payday_checklist"))


@main.route("/payday-checklist/preferences/<path:item_key>/unhide", methods=["POST"])
@login_required
def unhide_payday_checklist_item(item_key):
    pref = PaydayChecklistPreference.query.filter_by(item_key=item_key).first()
    if pref:
        pref.hidden = False
        audit("unhide_payday_checklist_item", "PaydayChecklistPreference", pref.label, "Restored to payday checklists")
        db.session.commit()
        flash("Checklist item restored.", "success")
    return redirect(url_for("main.payday_checklist"))


@main.route("/backup-restore", methods=["GET", "POST"])
@login_required
def backup_restore():
    db_path = get_database_path()
    last_backup = None
    backup_dir = os.path.join(current_app.instance_path, "backups")
    if os.path.isdir(backup_dir):
        backups = sorted([name for name in os.listdir(backup_dir) if name.endswith(".db")], reverse=True)
        last_backup = backups[0] if backups else None

    if request.method == "POST":
        upload = request.files.get("restore_file")
        confirm = request.form.get("confirm_restore") == "RESTORE"
        if not confirm:
            flash("Type RESTORE to confirm database restore.", "warning")
            return redirect(url_for("main.backup_restore"))
        if not upload or not upload.filename:
            flash("Choose a .db or .zip backup file.", "warning")
            return redirect(url_for("main.backup_restore"))
        if not check_upload_size(upload, MAX_RESTORE_UPLOAD_BYTES):
            flash("Backup file is too large. Maximum restore upload size is 50 MB.", "danger")
            return redirect(url_for("main.backup_restore"))
        restore_tmp_path = None
        try:
            os.makedirs(backup_dir, exist_ok=True)
            safety_backup = os.path.join(backup_dir, f"pre-restore-{datetime.now().strftime('%Y%m%d-%H%M%S')}.db")
            if os.path.exists(db_path):
                shutil.copy2(db_path, safety_backup)
            raw = upload.read()
            restore_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".db")
            restore_tmp_path = restore_tmp.name
            restore_tmp.close()
            if upload.filename.lower().endswith(".zip"):
                with ZipFile(io.BytesIO(raw), "r") as zipf:
                    db_names = [name for name in zipf.namelist() if name.endswith(".db")]
                    if not db_names:
                        raise ValueError("No .db file found inside ZIP backup.")
                    with open(restore_tmp.name, "wb") as out:
                        out.write(zipf.read(db_names[0]))
            elif upload.filename.lower().endswith(".db"):
                with open(restore_tmp.name, "wb") as out:
                    out.write(raw)
            else:
                raise ValueError("Upload a .db or .zip backup file.")
            db.session.remove()
            db.engine.dispose()
            shutil.copy2(restore_tmp.name, db_path)
            flash("Database restored. Restart Project Solace now so all pages use the restored file.", "success")
        except Exception as exc:
            flash(f"Restore failed: {exc}", "danger")
        finally:
            if restore_tmp_path:
                try:
                    os.unlink(restore_tmp_path)
                except OSError:
                    pass
        return redirect(url_for("main.backup_restore"))

    return render_template("backup_restore.html", db_path=db_path, last_backup=last_backup)


@main.route("/audit-log")
@login_required
def audit_log():
    rows = AuditLog.query.order_by(AuditLog.created_at.desc(), AuditLog.id.desc()).limit(200).all()
    return render_template("audit_log.html", rows=rows)


@main.route("/notifications", methods=["GET", "POST"])
@login_required
def notifications():
    settings_row = NotificationSetting.query.first()
    if not settings_row:
        settings_row = NotificationSetting()
        db.session.add(settings_row)
        db.session.commit()
    form = NotificationSettingsForm(obj=settings_row)
    if form.validate_on_submit():
        form.populate_obj(settings_row)
        db.session.commit()
        audit("update_notifications", "NotificationSetting", "Notifications", "Updated notification settings")
        flash("Notification settings saved.", "success")
        return redirect(url_for("main.notifications"))
    return render_template("notifications.html", form=form)


@main.route("/notifications/test", methods=["POST"])
@login_required
def test_notification():
    import urllib.request as _urllib_request
    import json as _json
    settings_row = NotificationSetting.query.first()
    if not settings_row or not settings_row.enabled:
        flash("Notifications are not enabled. Enable them and save first.", "warning")
        return redirect(url_for("main.notifications"))
    if settings_row.provider == "ntfy":
        url = (settings_row.webhook_url or "").strip()
        if not url:
            flash("No ntfy webhook URL configured.", "danger")
            return redirect(url_for("main.notifications"))
        try:
            req = _urllib_request.Request(url, data=b"Project Solace: test notification", method="POST")
            req.add_header("Title", "Test Notification")
            req.add_header("Tags", "bell")
            req.add_header("Content-Type", "text/plain")
            if settings_row.token:
                req.add_header("Authorization", f"Bearer {settings_row.token}")
            with _urllib_request.urlopen(req, timeout=8) as resp:
                if resp.status < 300:
                    flash("Test notification sent via ntfy.", "success")
                else:
                    flash(f"ntfy returned HTTP {resp.status}.", "danger")
        except Exception as exc:
            flash(f"ntfy error: {exc}", "danger")
    elif settings_row.provider == "Gotify/Webhook":
        url = (settings_row.webhook_url or "").strip()
        if not url:
            flash("No webhook URL configured.", "danger")
            return redirect(url_for("main.notifications"))
        try:
            payload = _json.dumps({"title": "Test Notification", "message": "Project Solace: test notification", "priority": 5}).encode()
            req = _urllib_request.Request(url, data=payload, method="POST")
            req.add_header("Content-Type", "application/json")
            if settings_row.token:
                req.add_header("X-Gotify-Key", settings_row.token)
            with _urllib_request.urlopen(req, timeout=8) as resp:
                if resp.status < 300:
                    flash("Test notification sent via webhook.", "success")
                else:
                    flash(f"Webhook returned HTTP {resp.status}.", "danger")
        except Exception as exc:
            flash(f"Webhook error: {exc}", "danger")
    else:
        flash("Select ntfy or Gotify/Webhook as provider first.", "warning")
    return redirect(url_for("main.notifications"))


@main.route("/cycle-history")
@login_required
def cycle_history():
    settings = get_settings()
    closeouts = CycleCloseout.query.order_by(CycleCloseout.cycle_start.desc()).all()
    rows = []
    for closeout in closeouts:
        try:
            cs = date.fromisoformat(closeout.cycle_start)
            ce = date.fromisoformat(closeout.cycle_end)
        except (ValueError, TypeError):
            cs = ce = None
        paid_total = 0
        unpaid_count = 0
        skipped_count = 0
        if cs and ce:
            occs = BillOccurrence.query.filter(
                BillOccurrence.due_date >= closeout.cycle_start,
                BillOccurrence.due_date <= closeout.cycle_end,
            ).all()
            paid_total = money(sum(o.amount for o in occs if o.status == "Paid"))
            unpaid_count = sum(1 for o in occs if o.status == "Upcoming")
            skipped_count = sum(1 for o in occs if o.status == "Skipped")
        rows.append({
            "closeout": closeout,
            "cs": cs,
            "ce": ce,
            "paid_total": paid_total,
            "unpaid_count": unpaid_count,
            "skipped_count": skipped_count,
        })
    return render_template("cycle_history.html", settings=settings, rows=rows)


@main.route("/annual-summary")
@login_required
def annual_summary():
    settings = get_settings()
    today = date.today()
    year_type = request.args.get("year_type", "calendar")
    if year_type == "financial":
        fy_year = today.year if today.month >= 7 else today.year - 1
        period_start = date(fy_year, 7, 1)
        period_end = date(fy_year + 1, 6, 30)
        period_label = f"FY {fy_year}/{str(fy_year + 1)[-2:]}"
    else:
        period_start = date(today.year, 1, 1)
        period_end = date(today.year, 12, 31)
        period_label = str(today.year)

    occurrences = BillOccurrence.query.filter(
        BillOccurrence.due_date >= period_start.isoformat(),
        BillOccurrence.due_date <= period_end.isoformat(),
    ).all()

    category_totals = {}
    for occ in occurrences:
        cat_name = occ.bill.category.name if occ.bill and occ.bill.category else "Uncategorised"
        if cat_name not in category_totals:
            category_totals[cat_name] = {"total": 0, "paid": 0, "unpaid": 0, "skipped": 0, "bills": {}}
        category_totals[cat_name]["total"] += occ.amount
        if occ.status == "Paid":
            category_totals[cat_name]["paid"] += occ.amount
        elif occ.status == "Skipped":
            category_totals[cat_name]["skipped"] += occ.amount
        else:
            category_totals[cat_name]["unpaid"] += occ.amount
        bill_name = occ.bill.name if occ.bill else "Unknown"
        category_totals[cat_name]["bills"][bill_name] = category_totals[cat_name]["bills"].get(bill_name, 0) + occ.amount

    categories = sorted([
        {
            "name": name,
            "total": money(data["total"]),
            "paid": money(data["paid"]),
            "unpaid": money(data["unpaid"]),
            "skipped": money(data["skipped"]),
            "bills": sorted([(k, money(v)) for k, v in data["bills"].items()], key=lambda x: -x[1]),
        }
        for name, data in category_totals.items()
    ], key=lambda x: -x["total"])

    grand_total = money(sum(c["total"] for c in categories))
    grand_paid = money(sum(c["paid"] for c in categories))

    return render_template(
        "annual_summary.html",
        settings=settings,
        categories=categories,
        grand_total=grand_total,
        grand_paid=grand_paid,
        period_label=period_label,
        period_start=period_start,
        period_end=period_end,
        year_type=year_type,
    )


def bills_to_rows():
    rows = []
    for bill in RecurringBill.query.order_by(RecurringBill.name).all():
        rows.append({
            "name": bill.name,
            "amount": bill.amount,
            "frequency": bill.frequency,
            "due_day": bill.due_day,
            "due_month": bill.due_month or "",
            "start_date": bill.start_date,
            "end_date": bill.end_date or "",
            "category": bill.category.name if bill.category else "",
            "active": "yes" if bill.active else "no",
            "autopay": "yes" if bill.autopay else "no",
            "account_name": bill.account_name or "",
            "include_in_set_aside": "yes" if bill.include_in_set_aside else "no",
            "notes": bill.notes or "",
        })
    return rows


def purchases_to_rows():
    rows = []
    for purchase in PlannedPurchase.query.order_by(PlannedPurchase.target_date).all():
        rows.append({
            "name": purchase.name,
            "target_amount": purchase.target_amount,
            "amount_saved": purchase.amount_saved,
            "target_date": purchase.target_date,
            "category": purchase.category.name if purchase.category else "",
            "priority": purchase.priority,
            "status": purchase.status,
            "purchase_scope": getattr(purchase, "purchase_scope", "Shared"),
            "owner_name": getattr(purchase, "owner_name", "") or "",
            "notes": purchase.notes or "",
        })
    return rows



def income_to_rows():
    rows = []
    for income in IncomeSource.query.order_by(IncomeSource.name).all():
        rows.append({
            "owner_name": getattr(income, "owner_name", "Household"),
            "name": income.name,
            "amount": income.amount,
            "frequency": income.frequency,
            "next_pay_date": income.next_pay_date,
            "active": "yes" if income.active else "no",
            "notes": income.notes or "",
        })
    return rows


def buckets_to_rows():
    rows = []
    for bucket in Bucket.query.order_by(Bucket.sort_order, Bucket.name).all():
        rows.append({
            "name": bucket.name,
            "percentage": bucket.percentage,
            "fixed_amount": bucket.fixed_amount if bucket.fixed_amount is not None else "",
            "rounding_increment": bucket.rounding_increment,
            "bucket_type": bucket.bucket_type,
            "cap_to_remaining": "yes" if getattr(bucket, "cap_to_remaining", False) else "no",
            "active": "yes" if bucket.active else "no",
            "sort_order": bucket.sort_order,
            "notes": bucket.notes or "",
        })
    return rows

def make_csv_response(filename, rows):
    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    else:
        output.write("")
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@main.route("/data")
@login_required
def data_tools():
    return render_template("data_tools.html", import_preview=load_bill_import_preview())


@main.route("/data/export/bills.csv")
@login_required
def export_bills_csv():
    return make_csv_response("project-solace-bills.csv", bills_to_rows())


@main.route("/data/export/purchases.csv")
@login_required
def export_purchases_csv():
    return make_csv_response("project-solace-planned-purchases.csv", purchases_to_rows())


@main.route("/data/export/income.csv")
@login_required
def export_income_csv():
    return make_csv_response("project-solace-income-sources.csv", income_to_rows())


@main.route("/data/export/buckets.csv")
@login_required
def export_buckets_csv():
    return make_csv_response("project-solace-buckets.csv", buckets_to_rows())


@main.route("/data/export/backup.xlsx")
@login_required
def export_backup_xlsx():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Bills"
    bill_rows = bills_to_rows()
    if bill_rows:
        ws.append(list(bill_rows[0].keys()))
        for row in bill_rows:
            ws.append(list(row.values()))
    ws2 = wb.create_sheet("Planned Purchases")
    purchase_rows = purchases_to_rows()
    if purchase_rows:
        ws2.append(list(purchase_rows[0].keys()))
        for row in purchase_rows:
            ws2.append(list(row.values()))
    ws3 = wb.create_sheet("Income Sources")
    income_rows = income_to_rows()
    if income_rows:
        ws3.append(list(income_rows[0].keys()))
        for row in income_rows:
            ws3.append(list(row.values()))
    ws4 = wb.create_sheet("Buckets")
    bucket_rows = buckets_to_rows()
    if bucket_rows:
        ws4.append(list(bucket_rows[0].keys()))
        for row in bucket_rows:
            ws4.append(list(row.values()))
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    audit("export_backup_xlsx", "Backup", "Readable XLSX backup", "Downloaded readable backup")
    db.session.commit()
    return send_temp_file(tmp.name, download_name="project-solace-backup.xlsx")


@main.route("/data/export/database.zip")
@login_required
def export_database_zip():
    db_path = current_app.config["SQLALCHEMY_DATABASE_URI"].replace("sqlite:///", "")
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    tmp.close()
    with ZipFile(tmp.name, "w", ZIP_DEFLATED) as zipf:
        if os.path.exists(db_path):
            zipf.write(db_path, arcname="solace.db")
    audit("export_database_zip", "Backup", "SQLite database ZIP", "Downloaded database backup")
    db.session.commit()
    return send_temp_file(tmp.name, download_name="project-solace-database-backup.zip")


def read_uploaded_rows(file_storage):
    filename = file_storage.filename.lower()
    raw = file_storage.read()
    if filename.endswith(".csv"):
        text = raw.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip().lower().replace(" ", "_") if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    item[headers[idx]] = value
            if any(value not in [None, ""] for value in item.values()):
                rows.append(item)
        return rows
    raise ValueError("Upload a CSV or XLSX file.")


def pick(row, *names, default=""):
    lowered = {str(k).lower().replace(" ", "_"): v for k, v in row.items()}
    for name in names:
        key = name.lower().replace(" ", "_")
        if key in lowered and lowered[key] not in [None, ""]:
            return lowered[key]
    return default



@main.route("/data/import/bills", methods=["POST"])
@login_required
def import_bills():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        flash("Choose a CSV or XLSX file first.", "warning")
        return redirect(url_for("main.data_tools"))
    if not check_upload_size(upload, MAX_IMPORT_UPLOAD_BYTES):
        flash("Import file is too large. Maximum import upload size is 5 MB.", "danger")
        return redirect(url_for("main.data_tools"))
    try:
        raw_rows = read_uploaded_rows(upload)
        if len(raw_rows) > MAX_IMPORT_PREVIEW_ROWS:
            raise ValueError(f"Import files are limited to {MAX_IMPORT_PREVIEW_ROWS} rows per preview.")
        parsed_rows, error_count = parse_bill_import_rows(raw_rows)
        save_bill_import_preview(parsed_rows)
        if error_count:
            flash(f"Import preview found {error_count} row(s) with errors. Fix the file or import only after reviewing.", "warning")
        else:
            flash(f"Import preview ready: {len(parsed_rows)} bill(s) detected.", "success")
        return redirect(url_for("main.data_tools"))
    except Exception as exc:
        flash(f"Import preview failed: {exc}", "danger")
        return redirect(url_for("main.data_tools"))


@main.route("/data/import/bills/confirm", methods=["POST"])
@login_required
def confirm_import_bills():
    parsed_rows = load_bill_import_preview()
    if not parsed_rows:
        flash("No bill import preview is waiting to be confirmed.", "warning")
        return redirect(url_for("main.data_tools"))
    imported = 0
    try:
        for row in parsed_rows:
            if row.get("errors"):
                continue
            category = get_or_create_category(row.get("category", ""), "Bill")
            bill = RecurringBill(
                name=row["name"],
                amount=float(row["amount"]),
                frequency=row["frequency"],
                due_day=int(row["due_day"]),
                due_month=row.get("due_month"),
                start_date=row["start_date"],
                end_date=row.get("end_date"),
                category_id=category.id if category else None,
                active=bool(row.get("active", True)),
                autopay=bool(row.get("autopay", False)),
                account_name=row.get("account_name", ""),
                include_in_set_aside=bool(row.get("include_in_set_aside", True)),
                notes=row.get("notes", ""),
            )
            db.session.add(bill)
            db.session.flush()
            regenerate_bill_occurrences(bill, scope="all_unpaid")
            imported += 1
        audit("import_bills", "RecurringBill", "Bill import", f"Imported {imported} bills from preview")
        db.session.commit()
        clear_bill_import_preview()
        flash(f"Imported {imported} recurring bills.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Import failed: {exc}", "danger")
    return redirect(url_for("main.data_tools"))


@main.route("/data/import/bills/cancel", methods=["POST"])
@login_required
def cancel_import_bills():
    clear_bill_import_preview()
    flash("Bill import preview cleared.", "info")
    return redirect(url_for("main.data_tools"))
